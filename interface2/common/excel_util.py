from pathlib import Path
from openpyxl import load_workbook
from common.exception_utils import exception_utils
import time
from common.text_util import read_txt_handel


@exception_utils
class ExcelUtil(object):

    def __init__(self, excel_path='%s/data/case_excel/接口测试框架实践用例.xlsx' % Path(__file__).parent.parent):
        self.wb = load_workbook(excel_path)
        self.base_dir = Path(__file__).parent.parent
        self.template = """{"id":0,"url":"","case_name":"","header":"","method":"","body":"",
        "expect":"","actual":"","valiadate":""},"""  # 这个是写入用例的模板

    @exception_utils
    def read_excel(self):
        """读取excel，处理数据，并返回一个格式处理后的字典"""
        value = []
        smoke_value = []
        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            cases_num = len(list(ws.values)) - 1  # 一个sheet中用例的数量
            case_list = list(ws.values)
            case_list.pop(0)  # 去掉表头
            cases_template = self.template * cases_num
            cases_template_list = eval("[" + cases_template[:-1] + "]")   # 与用例相同长度的模板

            for i in range(len(case_list)):  # i：第i个用例
                # 每个用例中字段是9个，因此这样写
                cases_template_list[i]['id'] = case_list[i][0]
                cases_template_list[i]['url'] = case_list[i][1]
                cases_template_list[i]['case_name'] = case_list[i][2]
                cases_template_list[i]['header'] = case_list[i][3]
                cases_template_list[i]['method'] = case_list[i][4]
                cases_template_list[i]['body'] = case_list[i][5]
                cases_template_list[i]['expect'] = case_list[i][6]
                cases_template_list[i]['actual'] = case_list[i][7]
                cases_template_list[i]['valiadate'] = case_list[i][8]
            value.append({"cases": cases_template_list})
        print(value)

        for v in value:
            for case in v['cases']:
                # print(case)
                if '正常' in str(case):
                    print(case['id'], case)
                    smoke_value.append(case)

        smoke = {"cases": smoke_value}
        return value, smoke

    @exception_utils
    def write_excel(self):
        """运行结果写入excel"""
        l_reponse, l_ispass = read_txt_handel()

        i = 0
        j = 0
        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            # 实际结果列
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8, min_col=8):
                for cell in row:
                    cell.value = l_reponse[i]
                    # print("resp:%s" % i, cell.value)
                    i += 1
            # 是否通过列
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=9, min_col=9):
                for cell in row:
                    cell.value = l_ispass[j]
                    # print("ispass%s:" % j, cell.value)
                    j += 1
        strftime = time.strftime("%Y%m%d_%H:%M:%S")
        save_path = "%s/output/run_result_excel/运行结果_%s.xlsx" % (self.base_dir, strftime)
        self.wb.save(save_path)
        return save_path


if __name__ == '__main__':
    excel_path = '/Users/dongshuai/PycharmProjects/interface2/data/case_excel/接口测试框架实践用例.xlsx'
    result_path = '/Users/dongshuai/PycharmProjects/interface2/output/run_result_ex' \
                  'cel/' + time.strftime("%Y%m%d_%H:%M:%S")
    ExcelUtil(excel_path).read_excel()
    # ExcelUtil(excel_path).write_excel()



